"""Download archived monthly STEO issues and extract the real-time residential
electricity price forecast (series ESRCUUS, nominal cents/kWh) from each vintage.

Source: https://www.eia.gov/outlooks/steo/archives/{mon}{yy}_base.xlsx
Each _base.xlsx is the STEO data workbook *as published* in that issue month, so
the ESRCUUS path it contains is a genuine real-time forecast/nowcast (no revision
leakage). We key each vintage by its issue month (from the filename) and store the
full monthly ESRCUUS path it published.

Output: raw/steo_vintages.json  (+ downloaded xlsx kept in raw/vintages/)
"""
from __future__ import annotations
import json, sys, urllib.request
from pathlib import Path

HERE = Path(__file__).resolve().parent
VDIR = HERE / "raw" / "vintages"
VDIR.mkdir(parents=True, exist_ok=True)
UA = {"User-Agent": "Mozilla/5.0 (research; CPI electricity STEO study)"}
MONS = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]


def add_months(ym: str, n: int) -> str:
    y, m = int(ym[:4]), int(ym[5:7])
    idx = (y * 12 + (m - 1)) + n
    return f"{idx // 12:04d}-{idx % 12 + 1:02d}"


def download(tag: str) -> Path | None:
    dest = VDIR / f"{tag}_base.xlsx"
    if dest.exists() and dest.stat().st_size > 50000:
        return dest
    url = f"https://www.eia.gov/outlooks/steo/archives/{tag}_base.xlsx"
    try:
        req = urllib.request.Request(url, headers=UA)
        data = urllib.request.urlopen(req, timeout=90).read()
    except Exception as exc:  # noqa: BLE001
        print(f"  {tag}: download failed ({type(exc).__name__})", flush=True)
        return None
    if len(data) < 50000:
        return None
    dest.write_bytes(data)
    return dest


def parse_vintage(path: Path, issue_month: str) -> dict | None:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # 1) Locate the ESRCUUS row by scanning column A of every sheet. Build each
    #    sheet's grid exactly once (read-only worksheets are single-pass).
    target_sheet = target_row = grid = None
    for sh in wb.sheetnames:
        g = list(wb[sh].iter_rows(values_only=True))  # 0-indexed list of tuples
        for ri, row in enumerate(g, start=1):
            if row and isinstance(row[0], str) and row[0].strip().upper() == "ESRCUUS":
                target_sheet, target_row, grid = sh, ri, g
                break
        if target_sheet:
            break
    if target_sheet is None:
        print(f"  {issue_month}: ESRCUUS row not found", flush=True)
        wb.close()
        return None

    # 2) Determine the table's first month. Primary source: the 'Dates' sheet
    #    "Table Beginning Month" (YYYYMM). Fallback: first year in the sheet's
    #    top header rows (STEO monthly tables always begin at January).
    begin_month = None
    if "Dates" in wb.sheetnames:
        for row in wb["Dates"].iter_rows(values_only=True):
            labels = [c for c in row if isinstance(c, str) and "beginning month" in c.lower()]
            if labels:
                for c in row:
                    if isinstance(c, (int, float)) and 190001 <= int(c) <= 210012:
                        iv = int(c)
                        begin_month = f"{iv // 100:04d}-{iv % 100:02d}"
                        break
            if begin_month:
                break

    def is_year(v):
        return isinstance(v, (int, float)) and 1990 <= v <= 2100 and float(v).is_integer()

    if begin_month is None:
        for cells in grid[:12]:
            years = [int(v) for v in cells if is_year(v)]
            if len(years) >= 2:
                begin_month = f"{min(years)}-01"
                break
    if begin_month is None:
        print(f"  {issue_month}: table beginning month not found", flush=True)
        wb.close()
        return None

    # 3) Read ESRCUUS values from the first numeric column (col A=id, B=label),
    #    mapping months sequentially from begin_month.
    val_row = grid[target_row - 1]
    start_col = None
    for i in range(2, len(val_row)):
        if isinstance(val_row[i], (int, float)):
            start_col = i
            break
    if start_col is None:
        wb.close()
        return None
    path_points = []
    m = begin_month
    for col in range(start_col, len(val_row)):
        v = val_row[col]
        if v is None or v == "":
            # STEO monthly tables are contiguous; stop at first trailing gap
            # but tolerate a single stray blank inside by breaking (values are dense).
            break
        try:
            fv = float(v)
        except (TypeError, ValueError):
            break
        path_points.append({"date": m, "value": fv})
        m = add_months(m, 1)

    wb.close()
    if not path_points:
        return None
    return {
        "issue_month": issue_month,
        "begin_month": begin_month,
        "sheet": target_sheet,
        "n_months": len(path_points),
        "first": path_points[0]["date"],
        "last": path_points[-1]["date"],
        "path": path_points,
    }


def main():
    args = sys.argv[1:]
    if args:  # optional explicit tags for testing, e.g. "jan15 jan20 jul26"
        tags = args
    else:
        tags = []
        for yy in range(15, 27):
            for mi, mon in enumerate(MONS, 1):
                if yy == 26 and mi > 7:  # today 2026-07; Aug+ not yet issued
                    continue
                tags.append(f"{mon}{yy:02d}")

    vintages = []
    for tag in tags:
        issue_month = f"20{tag[-2:]}-{MONS.index(tag[:3]) + 1:02d}"
        p = download(tag)
        if p is None:
            print(f"{tag}: MISSING", flush=True)
            continue
        v = parse_vintage(p, issue_month)
        if v is None:
            continue
        vintages.append(v)
        print(f"{tag} -> issue {issue_month}: {v['n_months']} mo [{v['first']}..{v['last']}] sheet={v['sheet']}", flush=True)

    vintages.sort(key=lambda x: x["issue_month"])
    out = HERE / "raw" / "steo_vintages.json"
    out.write_text(json.dumps({"series": "ESRCUUS", "unit": "cents/kWh",
                               "note": "Real-time STEO vintages; issue_month from archive filename.",
                               "vintages": vintages}, indent=1), encoding="utf-8")
    print(f"\nSaved {len(vintages)} vintages -> {out}", flush=True)


if __name__ == "__main__":
    main()
