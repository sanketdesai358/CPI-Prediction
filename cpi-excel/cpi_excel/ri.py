from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from .constants import BASE_URLS, RAW_DIR
from .labstat import normalize_item_name
from .sources import download_binary


@dataclass(frozen=True)
class RelativeImportanceData:
    ri_by_code: dict[str, float]
    weight_vintage: int | None
    december_month: date | None
    source_url: str
    unmatched_names: list[str]


ALIASES = {
    "food and beverages": "Food and beverages",
    "owners equivalent rent of residences": "Owners' equivalent rent of residences",
    "owners equivalent rent of primary residence": "Owners' equivalent rent of primary residence",
    "telephone hardware calculators and other consumer information items": "Telephone hardware, calculators, and other consumer information items",
}


def parse_ri_title(title: str) -> tuple[int | None, date | None]:
    vintage_match = re.search(r"\((\d{4})\s+Weights\)", title)
    month_match = re.search(
        r"(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})",
        title,
    )
    vintage = int(vintage_match.group(1)) if vintage_match else None
    dec_month = None
    if month_match:
        month_name, year = month_match.groups()
        month_number = {
            "January": 1,
            "February": 2,
            "March": 3,
            "April": 4,
            "May": 5,
            "June": 6,
            "July": 7,
            "August": 8,
            "September": 9,
            "October": 10,
            "November": 11,
            "December": 12,
        }[month_name]
        dec_month = date(int(year), month_number, 1)
    return vintage, dec_month


def load_current_relative_importance(
    items_df: pd.DataFrame, *, force: bool = False
) -> RelativeImportanceData:
    source_url = BASE_URLS["relative_importance_xlsx"]
    content = download_binary(
        source_url,
        RAW_DIR / "cpi-relative-importance.xlsx",
        force=force,
        timeout=120,
    )
    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = workbook["Table 1"] if "Table 1" in workbook.sheetnames else workbook.worksheets[0]

    title = str(sheet.cell(row=1, column=2).value or "")
    weight_vintage, december_month = parse_ri_title(title)

    name_to_code: dict[str, str] = {}
    for row in items_df.itertuples(index=False):
        name_to_code[normalize_item_name(str(row.item_name))] = str(row.item_code)

    ri_by_code: dict[str, float] = {}
    unmatched_names: list[str] = []

    for row in sheet.iter_rows(min_row=1, values_only=True):
        if len(row) < 3:
            continue
        raw_name = row[1]
        raw_value = row[2]
        if raw_name in (None, "") or raw_value in (None, ""):
            continue
        if isinstance(raw_name, str) and raw_name.strip().lower() == "item and group":
            continue
        try:
            value = float(raw_value)
        except (TypeError, ValueError):
            continue

        name = str(raw_name).strip()
        normalized = normalize_item_name(name)
        alias_name = ALIASES.get(normalized)
        if alias_name:
            normalized = normalize_item_name(alias_name)
        code = name_to_code.get(normalized)
        if not code:
            unmatched_names.append(name)
            continue
        ri_by_code[code] = value

    return RelativeImportanceData(
        ri_by_code=ri_by_code,
        weight_vintage=weight_vintage,
        december_month=december_month,
        source_url=source_url,
        unmatched_names=unmatched_names,
    )


def bls_ri_worked_example() -> dict[str, float]:
    """Published on the BLS relative-importance page, Table A."""
    return {
        "component_dec_ri": 7.039,
        "component_index_dec": 193.306,
        "component_index_month": 215.711,
        "all_items_index_dec": 241.432,
        "all_items_index_month": 246.819,
        "expected_updated_ri": 7.683,
    }
