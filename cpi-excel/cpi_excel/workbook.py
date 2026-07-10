from __future__ import annotations

import math
import os
import shutil
import subprocess
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .constants import BASE_URLS, FORMULA_ERROR_TOKENS, WORKBOOK_FILENAME
from .data import excel_date, month_display
from .math import one_month_contribution, price_updated_relative_importance
from .ri import RelativeImportanceData


BLUE = "1F4E79"
LIGHT_BLUE = "D9EAF7"
INPUT_BLUE = "1F4E79"
FORMULA_BLACK = "000000"
LINK_GREEN = "008000"
WHITE = "FFFFFF"
HEADER_FILL = PatternFill("solid", fgColor=BLUE)
SUBHEADER_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
INPUT_FONT = Font(name="Calibri", color=INPUT_BLUE)
FORMULA_FONT = Font(name="Calibri", color=FORMULA_BLACK)
LINK_FONT = Font(name="Calibri", color=LINK_GREEN, underline="single")
HEADER_FONT = Font(name="Calibri", color=WHITE, bold=True)
SUBHEADER_FONT = Font(name="Calibri", bold=True)
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=BLUE)
NEG_NUM_FMT = '0.000;[Red](0.000);-'
NEG_PCT_FMT = '0.0%;[Red](0.0%);-'
DATE_FMT = "mmm yyyy"


@dataclass(frozen=True)
class WorkbookBuildInfo:
    path: Path
    latest_month: pd.Timestamp
    latest_months: list[pd.Timestamp]
    formula_error_literals: list[str]
    libreoffice_checked: bool


def _set_title(ws: openpyxl.worksheet.worksheet.Worksheet, title: str, end_col: int = 8) -> None:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if end_col > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)


def _style_header_row(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, max_col: int) -> None:
    for cell in ws[row][0:max_col]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_subheader_row(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, max_col: int) -> None:
    for cell in ws[row][0:max_col]:
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autofit_basic(ws: openpyxl.worksheet.worksheet.Worksheet, *, max_width: int = 48) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            value = str(cell.value)
            widths[cell.column] = min(max(widths.get(cell.column, 0), len(value) + 2), max_width)
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = max(10, width)


def _apply_table_settings(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int = 1) -> None:
    if ws.max_row >= header_row and ws.max_column >= 1:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = f"A{header_row + 1}"


def _metadata_comment(url: str, date_label: str | None = None) -> Comment:
    label = date_label or datetime.utcnow().strftime("%Y-%m-%d")
    return Comment(f"Source: BLS page, {label}, {url}", "Codex")


def _write_link(cell: openpyxl.cell.cell.Cell, label: str, url: str) -> None:
    cell.value = label
    cell.hyperlink = url
    cell.font = LINK_FONT


def _idx_formula(sheet_name: str, month_expr: str, series_expr: str, last_col: str) -> str:
    return (
        f"INDEX('{sheet_name}'!$A:${last_col},"
        f"MATCH({month_expr},'{sheet_name}'!$A:$A,0),"
        f"MATCH({series_expr},'{sheet_name}'!$1:$1,0))"
    )


def _safe(formula_body: str) -> str:
    return f'=IFERROR({formula_body},"")'


def _write_dataframe_sheet(
    wb: Workbook,
    sheet_name: str,
    data: pd.DataFrame,
    series_label: str,
) -> None:
    ws = wb.create_sheet(sheet_name)
    headers = ["Ref Month", *[str(column) for column in data.columns if column != "ref_month"]]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for cell in ws[1][1:]:
        url = f"https://data.bls.gov/timeseries/{cell.value}"
        cell.hyperlink = url
        cell.font = LINK_FONT
        cell.comment = _metadata_comment(url)

    for _, row in data.iterrows():
        values = []
        for column in data.columns:
            value = row[column]
            if pd.isna(value):
                value = None
            values.append(value)
        ws.append(values)

    for row in ws.iter_rows(min_row=2):
        row[0].number_format = DATE_FMT
        row[0].font = INPUT_FONT
        for cell in row[1:]:
            cell.number_format = NEG_NUM_FMT
            cell.font = INPUT_FONT
    ws["A1"].comment = _metadata_comment(BASE_URLS["bls_api"])
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    _apply_table_settings(ws, 1)
    _autofit_basic(ws, max_width=18)
    ws.column_dimensions["A"].width = 12
    ws["A1"].value = f"Ref Month ({series_label})"


def _write_readme(
    wb: Workbook,
    latest_months: list[pd.Timestamp],
    release_text: str,
    api_key_present: bool,
) -> None:
    ws = wb.active
    ws.title = "README"
    _set_title(ws, "CPI Component Workbook", 4)
    rows = [
        ("Build timestamp", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
        ("Data as of", month_display(latest_months[-1])),
        ("Latest 3 reference months", ", ".join(month_display(month) for month in latest_months)),
        ("Next CPI release", release_text or "See BLS release calendar"),
        ("BLS_API_KEY detected", "yes" if api_key_present else "no - using unregistered API limits/fallbacks"),
        ("Refresh command", "BLS_API_KEY=... python build_workbook.py"),
        ("Source of truth", "BLS Public Data API, LABSTAT cu.item/cu.series, and BLS relative-importance tables"),
    ]
    for row_index, (label, value) in enumerate(rows, start=3):
        ws.cell(row=row_index, column=1, value=label).font = SUBHEADER_FONT
        ws.cell(row=row_index, column=2, value=value).font = INPUT_FONT
    ws["A12"] = "Notes"
    ws["A12"].font = SUBHEADER_FONT
    ws["A13"] = (
        "Derived m/m, y/y, price-updated relative importance, and contribution cells are Excel formulas "
        "that reference Data_NSA and Data_SA."
    )
    ws["A13"].alignment = Alignment(wrap_text=True)
    ws["A15"] = "Primary BLS URLs"
    ws["A15"].font = SUBHEADER_FONT
    for row_index, (key, url) in enumerate(BASE_URLS.items(), start=16):
        ws.cell(row=row_index, column=1, value=key)
        _write_link(ws.cell(row=row_index, column=2), url, url)
    _autofit_basic(ws)
    ws.column_dimensions["B"].width = 80


def _write_component_tree(wb: Workbook, registry: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Component_Tree")
    headers = ["item_code", "name", "parent", "display_level", "is_item_stratum", "sort_sequence"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for entry in registry:
        ws.append(
            [
                entry["item_code"],
                entry["name"],
                entry.get("parent"),
                entry["display_level"],
                entry["is_item_stratum"],
                entry["sort_sequence"],
            ]
        )
        ws.cell(ws.max_row, 2).alignment = Alignment(indent=min(int(entry["display_level"]), 12))
    _apply_table_settings(ws, 1)
    _autofit_basic(ws)


def _write_methodology(wb: Workbook, registry: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Methodology")
    headers = [
        "item_code",
        "component",
        "formula",
        "collection",
        "qa_method",
        "alt_data",
        "factsheet_or_index",
        "handbook_calculation",
        "nsa_series",
        "sa_series",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for entry in registry:
        row = ws.max_row + 1
        ws.append(
            [
                entry["item_code"],
                entry["name"],
                entry["formula"],
                entry["collection"],
                entry["qa_method"],
                entry["alt_data"],
                "BLS methodology",
                "BLS Handbook",
                entry["series_nsa"],
                entry.get("series_sa") or "",
            ]
        )
        for col in range(3, 7):
            ws.cell(row=row, column=col).comment = _metadata_comment(BASE_URLS["factsheets"])
            ws.cell(row=row, column=col).font = INPUT_FONT
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        _write_link(ws.cell(row=row, column=7), "factsheet/index", entry["links"][0])
        _write_link(ws.cell(row=row, column=8), "calculation handbook", BASE_URLS["handbook_calculation"])
        _write_link(
            ws.cell(row=row, column=9),
            entry["series_nsa"],
            f"https://data.bls.gov/timeseries/{entry['series_nsa']}",
        )
        if entry.get("series_sa"):
            _write_link(
                ws.cell(row=row, column=10),
                entry["series_sa"],
                f"https://data.bls.gov/timeseries/{entry['series_sa']}",
            )
    _apply_table_settings(ws, 1)
    _autofit_basic(ws)
    ws.column_dimensions["D"].width = 42
    ws.column_dimensions["E"].width = 48


def _write_weights(
    wb: Workbook,
    registry: list[dict[str, Any]],
    ri_data: RelativeImportanceData,
    latest_months: list[pd.Timestamp],
    nsa_last_col: str,
) -> None:
    ws = wb.create_sheet("Weights")
    headers = [
        "item_code",
        "component",
        "series_nsa",
        "December RI %",
        "December RI month",
        "CE weight vintage",
    ]
    month_start_col = len(headers) + 1
    headers.extend([f"Price-updated RI {month_display(month)}" for month in latest_months])
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    dec_date = ri_data.december_month
    source_date = dec_date.isoformat() if dec_date else None
    for entry in registry:
        row = ws.max_row + 1
        dec_ri = ri_data.ri_by_code.get(entry["item_code"])
        ws.append(
            [
                entry["item_code"],
                entry["name"],
                entry["series_nsa"],
                dec_ri,
                dec_date,
                ri_data.weight_vintage,
            ]
        )
        for col in (1, 2, 3, 4, 5, 6):
            ws.cell(row=row, column=col).font = INPUT_FONT
        ws.cell(row=row, column=4).number_format = NEG_NUM_FMT
        ws.cell(row=row, column=4).comment = _metadata_comment(ri_data.source_url, source_date)
        ws.cell(row=row, column=5).number_format = DATE_FMT
        ws.cell(row=row, column=5).comment = _metadata_comment(ri_data.source_url, source_date)
        ws.cell(row=row, column=6).comment = _metadata_comment(BASE_URLS["relative_importance"], source_date)

        for offset, month in enumerate(latest_months):
            col = month_start_col + offset
            header_cell = f"{get_column_letter(col)}$1"
            component_month = _idx_formula("Data_NSA", header_cell, f"$C{row}", nsa_last_col)
            component_dec = _idx_formula("Data_NSA", f"$E{row}", f"$C{row}", nsa_last_col)
            all_month = _idx_formula("Data_NSA", header_cell, '"CUUR0000SA0"', nsa_last_col)
            all_dec = _idx_formula("Data_NSA", f"$E{row}", '"CUUR0000SA0"', nsa_last_col)
            ws.cell(row=row, column=col).value = _safe(
                f"100*($D{row}*({component_month}/{component_dec}))/(100*({all_month}/{all_dec}))"
            )
            ws.cell(row=row, column=col).font = FORMULA_FONT
            ws.cell(row=row, column=col).number_format = NEG_NUM_FMT

    for offset, month in enumerate(latest_months):
        cell = ws.cell(row=1, column=month_start_col + offset)
        cell.value = excel_date(month)
        cell.number_format = DATE_FMT
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    _apply_table_settings(ws, 1)
    _autofit_basic(ws)


def _write_series_catalog(wb: Workbook, registry: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Series_Catalog")
    headers = ["series_id", "item_code", "component", "seasonal", "data_url"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for entry in registry:
        for seasonal, series_id in (("NSA", entry["series_nsa"]), ("SA", entry.get("series_sa"))):
            if not series_id:
                continue
            row = ws.max_row + 1
            url = f"https://data.bls.gov/timeseries/{series_id}"
            ws.append([series_id, entry["item_code"], entry["name"], seasonal, url])
            _write_link(ws.cell(row=row, column=1), series_id, url)
            _write_link(ws.cell(row=row, column=5), url, url)
    _apply_table_settings(ws, 1)
    _autofit_basic(ws)


def _prior_ri_formula(
    row: int,
    month_cell: str,
    nsa_last_col: str,
) -> str:
    dec_ri = f"INDEX('Weights'!$D:$D,MATCH($A{row},'Weights'!$A:$A,0))"
    dec_month = f"INDEX('Weights'!$E:$E,MATCH($A{row},'Weights'!$A:$A,0))"
    component_prior = _idx_formula("Data_NSA", f"EDATE({month_cell},-1)", f"$D{row}", nsa_last_col)
    component_dec = _idx_formula("Data_NSA", dec_month, f"$D{row}", nsa_last_col)
    all_prior = _idx_formula("Data_NSA", f"EDATE({month_cell},-1)", '"CUUR0000SA0"', nsa_last_col)
    all_dec = _idx_formula("Data_NSA", dec_month, '"CUUR0000SA0"', nsa_last_col)
    return f"100*({dec_ri}*({component_prior}/{component_dec}))/(100*({all_prior}/{all_dec}))"


def _write_latest_3m(
    wb: Workbook,
    registry: list[dict[str, Any]],
    latest_months: list[pd.Timestamp],
    nsa_last_col: str,
    sa_last_col: str,
) -> dict[str, int]:
    ws = wb.create_sheet("Latest_3M")
    _set_title(ws, "Latest 3 Published Months", 12)
    id_headers = [
        "item_code",
        "component",
        "parent",
        "series_nsa",
        "series_sa",
        "formula",
        "collection",
        "Dec RI %",
    ]
    for col, header in enumerate(id_headers, start=1):
        ws.cell(row=3, column=col, value=header)
    month_start = len(id_headers) + 1
    month_width = 6
    subheaders = ["NSA index", "SA index", "SA m/m %", "NSA y/y %", "RI weight", "SA contrib pp"]
    for month_index, month in enumerate(latest_months):
        start_col = month_start + month_index * month_width
        end_col = start_col + month_width - 1
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        cell = ws.cell(row=2, column=start_col, value=excel_date(month))
        cell.number_format = DATE_FMT
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        for offset, subheader in enumerate(subheaders):
            ws.cell(row=3, column=start_col + offset, value=subheader)
    _style_header_row(ws, 3, month_start + len(latest_months) * month_width - 1)
    for col in range(1, month_start):
        ws.cell(row=2, column=col).fill = HEADER_FILL

    month_col_for_latest: dict[str, int] = {}
    for entry in registry:
        row = ws.max_row + 1
        ws.append(
            [
                entry["item_code"],
                entry["name"],
                entry.get("parent"),
                entry["series_nsa"],
                entry.get("series_sa") or "",
                entry["formula"],
                entry["collection"],
                f'=IFERROR(INDEX(Weights!$D:$D,MATCH($A{row},Weights!$A:$A,0)),"")',
            ]
        )
        for col in range(1, 8):
            ws.cell(row=row, column=col).font = INPUT_FONT
        ws.cell(row=row, column=8).font = FORMULA_FONT
        ws.cell(row=row, column=8).number_format = NEG_NUM_FMT

        for month_index, month in enumerate(latest_months):
            start_col = month_start + month_index * month_width
            month_cell = f"{get_column_letter(start_col)}$2"
            nsa_current = _idx_formula("Data_NSA", month_cell, f"$D{row}", nsa_last_col)
            nsa_lag = _idx_formula("Data_NSA", f"EDATE({month_cell},-12)", f"$D{row}", nsa_last_col)
            sa_current = _idx_formula("Data_SA", month_cell, f"$E{row}", sa_last_col)
            sa_prior = _idx_formula("Data_SA", f"EDATE({month_cell},-1)", f"$E{row}", sa_last_col)

            formulas = [
                _safe(nsa_current),
                f'=IF($E{row}="","",{_safe(sa_current)[1:]})',
                f'=IF($E{row}="","",{_safe(f"{sa_current}/{sa_prior}-1")[1:]})',
                _safe(f"{nsa_current}/{nsa_lag}-1"),
                _safe(
                    f"INDEX('Weights'!$G:$I,MATCH($A{row},'Weights'!$A:$A,0),"
                    f"MATCH({month_cell},'Weights'!$G$1:$I$1,0))"
                ),
                f'=IF($E{row}="","",{_safe(f"{_prior_ri_formula(row, month_cell, nsa_last_col)}*({sa_current}/{sa_prior}-1)")[1:]})',
            ]
            for offset, formula in enumerate(formulas):
                cell = ws.cell(row=row, column=start_col + offset, value=formula)
                cell.font = FORMULA_FONT
                cell.number_format = NEG_NUM_FMT
            ws.cell(row=row, column=start_col + 2).number_format = NEG_PCT_FMT
            ws.cell(row=row, column=start_col + 3).number_format = NEG_PCT_FMT
            month_col_for_latest[month.strftime("%Y-%m")] = start_col

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(ws.max_column)}{ws.max_row}"
    _autofit_basic(ws, max_width=32)
    ws.column_dimensions["B"].width = 44
    return month_col_for_latest


def _value_at(wide: pd.DataFrame, series_id: str, month: pd.Timestamp) -> float | None:
    if series_id not in wide.columns:
        return None
    matched = wide.loc[wide["ref_month"] == month, series_id]
    if matched.empty or pd.isna(matched.iloc[0]):
        return None
    return float(matched.iloc[0])


def _top_contributors(
    registry: list[dict[str, Any]],
    ri_data: RelativeImportanceData,
    wide_nsa: pd.DataFrame,
    wide_sa: pd.DataFrame,
    latest_month: pd.Timestamp,
    count: int = 10,
) -> list[dict[str, Any]]:
    if ri_data.december_month is None:
        return []
    dec_month = pd.Timestamp(ri_data.december_month)
    prior_month = latest_month - pd.DateOffset(months=1)
    all_prior = _value_at(wide_nsa, "CUUR0000SA0", prior_month)
    all_dec = _value_at(wide_nsa, "CUUR0000SA0", dec_month)
    if not all_prior or not all_dec:
        return []

    rows: list[dict[str, Any]] = []
    for entry in registry:
        if not entry.get("is_item_stratum") or not entry.get("series_sa"):
            continue
        dec_ri = ri_data.ri_by_code.get(entry["item_code"])
        if dec_ri is None:
            continue
        comp_prior_nsa = _value_at(wide_nsa, entry["series_nsa"], prior_month)
        comp_dec_nsa = _value_at(wide_nsa, entry["series_nsa"], dec_month)
        current_sa = _value_at(wide_sa, entry["series_sa"], latest_month)
        prior_sa = _value_at(wide_sa, entry["series_sa"], prior_month)
        if not all([comp_prior_nsa, comp_dec_nsa, current_sa, prior_sa]):
            continue
        try:
            prior_ri = price_updated_relative_importance(
                dec_ri, comp_prior_nsa, comp_dec_nsa, all_prior, all_dec
            )
            contribution = one_month_contribution(prior_ri, current_sa, prior_sa)
        except ZeroDivisionError:
            continue
        if math.isfinite(contribution):
            rows.append(
                {
                    "item_code": entry["item_code"],
                    "name": entry["name"],
                    "contribution": contribution,
                }
            )
    return sorted(rows, key=lambda row: abs(row["contribution"]), reverse=True)[:count]


def _latest_sheet_col(month_col_map: dict[str, int], month: pd.Timestamp, offset: int) -> str:
    return get_column_letter(month_col_map[month.strftime("%Y-%m")] + offset)


def _write_dashboard(
    wb: Workbook,
    latest_months: list[pd.Timestamp],
    top_contributors: list[dict[str, Any]],
    month_col_map: dict[str, int],
) -> None:
    ws = wb.create_sheet("Dashboard", 1)
    _set_title(ws, "CPI Dashboard", 10)
    ws["A3"] = "Data as of"
    ws["B3"] = month_display(latest_months[-1])
    ws["A5"] = "Headline and core"
    ws["A5"].font = SUBHEADER_FONT
    headers = ["Metric", *[month_display(month) for month in latest_months]]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=6, column=col, value=header)
    _style_header_row(ws, 6, len(headers))
    metrics = [
        ("All items SA m/m", "SA0", 2, NEG_PCT_FMT),
        ("All items NSA y/y", "SA0", 3, NEG_PCT_FMT),
        ("Core SA m/m", "SA0L1E", 2, NEG_PCT_FMT),
        ("Core NSA y/y", "SA0L1E", 3, NEG_PCT_FMT),
    ]
    for row_offset, (label, code, value_offset, number_format) in enumerate(metrics, start=7):
        ws.cell(row=row_offset, column=1, value=label)
        for month_index, month in enumerate(latest_months, start=2):
            col_letter = _latest_sheet_col(month_col_map, month, value_offset)
            formula = f'=IFERROR(INDEX(Latest_3M!${col_letter}:${col_letter},MATCH("{code}",Latest_3M!$A:$A,0)),"")'
            cell = ws.cell(row=row_offset, column=month_index, value=formula)
            cell.number_format = number_format
            cell.font = FORMULA_FONT

    latest = latest_months[-1]
    ws["A13"] = f"Top contributors, {month_display(latest)}"
    ws["A13"].font = SUBHEADER_FONT
    for col, header in enumerate(["Rank", "item_code", "component", "SA contribution pp"], start=1):
        ws.cell(row=14, column=col, value=header)
    _style_header_row(ws, 14, 4)
    latest_contrib_col = _latest_sheet_col(month_col_map, latest, 5)
    for rank, row_data in enumerate(top_contributors, start=1):
        row = 14 + rank
        code = row_data["item_code"]
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=code)
        ws.cell(row=row, column=3, value=row_data["name"])
        cell = ws.cell(
            row=row,
            column=4,
            value=f'=IFERROR(INDEX(Latest_3M!${latest_contrib_col}:${latest_contrib_col},MATCH(B{row},Latest_3M!$A:$A,0)),"")',
        )
        cell.number_format = NEG_NUM_FMT
        cell.font = FORMULA_FONT

    ws["F5"] = "Summary groups"
    ws["F5"].font = SUBHEADER_FONT
    for col, header in enumerate(["Group", "Latest SA m/m", "Contribution pp"], start=6):
        ws.cell(row=6, column=col, value=header)
    _style_header_row(ws, 6, 8)
    groups = [("Food", "SAF1"), ("Energy", "SA0E"), ("Shelter", "SAH1"), ("Core", "SA0L1E")]
    latest_mm_col = _latest_sheet_col(month_col_map, latest, 2)
    latest_group_contrib_col = _latest_sheet_col(month_col_map, latest, 5)
    for row_index, (label, code) in enumerate(groups, start=7):
        ws.cell(row=row_index, column=6, value=label)
        ws.cell(
            row=row_index,
            column=7,
            value=f'=IFERROR(INDEX(Latest_3M!${latest_mm_col}:${latest_mm_col},MATCH("{code}",Latest_3M!$A:$A,0)),"")',
        ).number_format = NEG_PCT_FMT
        ws.cell(
            row=row_index,
            column=8,
            value=f'=IFERROR(INDEX(Latest_3M!${latest_group_contrib_col}:${latest_group_contrib_col},MATCH("{code}",Latest_3M!$A:$A,0)),"")',
        ).number_format = NEG_NUM_FMT

    _autofit_basic(ws, max_width=36)


def _write_sources(wb: Workbook) -> None:
    ws = wb.create_sheet("Sources")
    ws.append(["source_key", "url", "use"])
    _style_header_row(ws, 1, 3)
    uses = {
        "bls_api": "Index levels, NSA and SA, all series",
        "api_registration": "API key registration and limits",
        "labstat_base": "Flat-file CPI metadata and bulk data",
        "relative_importance": "Annual and monthly RI methodology",
        "news_release": "Latest CPI release text",
        "supplemental": "Machine-readable release companion files",
        "release_calendar": "CPI release dates",
        "handbook_calculation": "CPI formula and calculation methods",
        "factsheets": "Per-component methodology factsheets",
    }
    for key, url in BASE_URLS.items():
        row = ws.max_row + 1
        ws.append([key, url, uses.get(key, "")])
        _write_link(ws.cell(row=row, column=2), url, url)
    _apply_table_settings(ws, 1)
    _autofit_basic(ws)
    ws.column_dimensions["B"].width = 82


def scan_formula_error_literals(path: Path) -> list[str]:
    wb = load_workbook(path, data_only=False, read_only=True)
    hits: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                if any(token in value for token in FORMULA_ERROR_TOKENS):
                    hits.append(f"{ws.title}!{cell.coordinate}: {value[:80]}")
    return hits


def recalculate_with_libreoffice(path: Path) -> bool:
    executable = shutil.which("soffice") or shutil.which("libreoffice")
    if not executable:
        return False
    output_dir = path.parent / "recalc"
    output_dir.mkdir(exist_ok=True)
    subprocess.run(
        [
            executable,
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(output_dir),
            str(path),
        ],
        check=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    recalculated = output_dir / path.name
    if recalculated.exists():
        shutil.copy2(recalculated, path)
    return True


def build_workbook(
    output_path: Path,
    registry: list[dict[str, Any]],
    ri_data: RelativeImportanceData,
    wide_nsa: pd.DataFrame,
    wide_sa: pd.DataFrame,
    latest_months: list[pd.Timestamp],
    release_text: str,
) -> WorkbookBuildInfo:
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    _write_readme(wb, latest_months, release_text, bool(os.environ.get("BLS_API_KEY")))
    _write_component_tree(wb, registry)
    _write_methodology(wb, registry)

    nsa_series_cols = [column for column in wide_nsa.columns if column != "ref_month"]
    sa_series_cols = [column for column in wide_sa.columns if column != "ref_month"]
    nsa_last_col = get_column_letter(len(nsa_series_cols) + 1)
    sa_last_col = get_column_letter(len(sa_series_cols) + 1)

    _write_weights(wb, registry, ri_data, latest_months, nsa_last_col)
    month_col_map = _write_latest_3m(wb, registry, latest_months, nsa_last_col, sa_last_col)
    _write_series_catalog(wb, registry)
    _write_dataframe_sheet(wb, "Data_NSA", wide_nsa, "NSA")
    _write_dataframe_sheet(wb, "Data_SA", wide_sa, "SA")
    _write_sources(wb)

    top = _top_contributors(registry, ri_data, wide_nsa, wide_sa, latest_months[-1])
    _write_dashboard(wb, latest_months, top, month_col_map)

    desired_order = [
        "README",
        "Dashboard",
        "Component_Tree",
        "Methodology",
        "Weights",
        "Latest_3M",
        "Series_Catalog",
        "Data_NSA",
        "Data_SA",
        "Sources",
    ]
    wb._sheets = [wb[sheet] for sheet in desired_order if sheet in wb.sheetnames]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    libreoffice_checked = recalculate_with_libreoffice(output_path)
    errors = scan_formula_error_literals(output_path)
    return WorkbookBuildInfo(
        path=output_path,
        latest_month=latest_months[-1],
        latest_months=latest_months,
        formula_error_literals=errors,
        libreoffice_checked=libreoffice_checked,
    )
