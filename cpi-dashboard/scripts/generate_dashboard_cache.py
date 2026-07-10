from __future__ import annotations

import json
import math
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
WORKSPACE = ROOT.parent
EXCEL_REPO = WORKSPACE / "cpi-excel"
WORKBOOK = EXCEL_REPO / "cpi_component_workbook.xlsx"
REGISTRY = ROOT / "src" / "data" / "registry.json"
OUTPUT = ROOT / "src" / "data" / "dashboard-cache.json"
CALENDAR_OUTPUT = ROOT / "src" / "data" / "release-calendar.json"


def month_key(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, date):
        return value.strftime("%Y-%m")
    text = str(value)
    match = re.search(r"(\d{4})-(\d{2})", text)
    if match:
        return f"{match.group(1)}-{match.group(2)}"
    return None


def parse_month(key: str) -> date:
    year, month = key.split("-")
    return date(int(year), int(month), 1)


def add_months(key: str, months: int) -> str:
    current = parse_month(key)
    index = current.year * 12 + current.month - 1 + months
    return f"{index // 12:04d}-{index % 12 + 1:02d}"


def read_wide_sheet(wb: openpyxl.Workbook, sheet_name: str) -> dict[str, dict[str, float | None]]:
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    series = [str(value) for value in headers[1:]]
    data: dict[str, dict[str, float | None]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = month_key(row[0])
        if not key:
            continue
        values: dict[str, float | None] = {}
        for series_id, raw in zip(series, row[1:]):
            try:
                values[series_id] = float(raw) if raw not in (None, "") else None
            except (TypeError, ValueError):
                values[series_id] = None
        data[key] = values
    return data


def read_weights(wb: openpyxl.Workbook) -> tuple[dict[str, dict[str, Any]], int | None, str | None]:
    ws = wb["Weights"]
    weights: dict[str, dict[str, Any]] = {}
    vintage = None
    december = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[0]
        if not code:
            continue
        dec_month = month_key(row[4])
        weights[str(code)] = {
            "decRi": float(row[3]) if row[3] not in (None, "") else None,
            "decemberMonth": dec_month,
            "weightVintage": int(row[5]) if row[5] not in (None, "") else None,
        }
        if vintage is None and weights[str(code)]["weightVintage"]:
            vintage = weights[str(code)]["weightVintage"]
        if december is None and dec_month:
            december = dec_month
    return weights, vintage, december


def val(data: dict[str, dict[str, float | None]], month: str, series_id: str | None) -> float | None:
    if not series_id:
        return None
    return data.get(month, {}).get(series_id)


def pct_change(current: float | None, prior: float | None) -> float | None:
    if current is None or prior in (None, 0):
        return None
    return current / prior - 1.0


def safe_mean(values: list[float | None]) -> float | None:
    clean = [float(value) for value in values if value is not None and math.isfinite(float(value))]
    return sum(clean) / len(clean) if clean else None


def derived_sa_proxy(
    data: dict[str, dict[str, float | None]],
    months: list[str],
    series_id: str | None,
) -> tuple[dict[str, float | None], dict[str, float | None]]:
    """Build a lightweight SA proxy when BLS publishes NSA but no SA series.

    This does not replace official CUSR data. It removes the component's
    average calendar-month seasonal pattern from NSA m/m changes, then chains a
    display index from the first available NSA level so tables/charts can show a
    continuous SA-style movement.
    """
    if not series_id:
        return {}, {}
    raw_by_month: dict[str, float | None] = {}
    for month in months:
        raw_by_month[month] = pct_change(val(data, month, series_id), val(data, add_months(month, -1), series_id))
    overall = safe_mean(list(raw_by_month.values()))
    if overall is None:
        return {}, {}
    seasonal: dict[str, float] = {}
    for raw_month in range(1, 13):
        key = f"{raw_month:02d}"
        month_mean = safe_mean([value for month, value in raw_by_month.items() if month[-2:] == key])
        seasonal[key] = 0.0 if month_mean is None else month_mean - overall
    proxy_mm: dict[str, float | None] = {}
    proxy_index: dict[str, float | None] = {}
    for month in months:
        raw = raw_by_month.get(month)
        proxy_mm[month] = None if raw is None else raw - seasonal[month[-2:]]
        prior = add_months(month, -1)
        nsa_index = val(data, month, series_id)
        if proxy_mm[month] is None:
            proxy_index[month] = None
        elif proxy_index.get(prior) not in (None, 0):
            proxy_index[month] = proxy_index[prior] * (1.0 + proxy_mm[month])
        else:
            proxy_index[month] = nsa_index
    return proxy_index, proxy_mm


def price_updated_ri(
    dec_ri: float | None,
    component_month: float | None,
    component_dec: float | None,
    all_month: float | None,
    all_dec: float | None,
) -> float | None:
    if None in (dec_ri, component_month, component_dec, all_month, all_dec):
        return None
    if component_dec == 0 or all_dec == 0 or all_month == 0:
        return None
    return 100.0 * (dec_ri * (component_month / component_dec)) / (100.0 * (all_month / all_dec))


def contribution_pp(prior_ri: float | None, current: float | None, prior: float | None) -> float | None:
    change = pct_change(current, prior)
    if prior_ri is None or change is None:
        return None
    return prior_ri * change


def parse_release_calendar() -> list[dict[str, str]]:
    path = EXCEL_REPO / "data" / "raw" / "release-calendar-cpi.html"
    if not path.exists():
        return [{"releaseDate": "2026-07-14", "releaseTime": "8:30 a.m. ET", "text": "Consumer Price Index"}]
    html = path.read_text(encoding="utf-8", errors="ignore")
    entries: list[dict[str, str]] = []
    for match in re.finditer(
        r"<tr[^>]*>\s*<td>(.*?)</td>\s*<td>(.*?)</td>\s*<td>(.*?)</td>\s*</tr>",
        html,
        flags=re.S,
    ):
        reference_month = re.sub(r"<[^>]+>", "", match.group(1)).strip()
        release_date_text = re.sub(r"<[^>]+>", "", match.group(2)).strip()
        time_text = re.sub(r"<[^>]+>", "", match.group(3)).strip()
        date_match = re.match(r"^([A-Z][a-z]+)\.?\s+(\d{1,2}),\s+(\d{4})$", release_date_text)
        if not date_match:
            continue
        month, day, year = date_match.groups()
        try:
            release = datetime.strptime(f"{month} {day} {year}", "%B %d %Y").date()
        except ValueError:
            try:
                release = datetime.strptime(f"{month}. {day} {year}", "%b. %d %Y").date()
            except ValueError:
                continue
        entries.append(
            {
                "releaseDate": release.isoformat(),
                "releaseTime": re.sub(r"\s+", " ", time_text).strip(),
                "text": f"Consumer Price Index: {reference_month}",
            }
        )
    if not entries:
        entries.append({"releaseDate": "2026-07-14", "releaseTime": "8:30 a.m. ET", "text": "Consumer Price Index"})
    return sorted({entry["releaseDate"]: entry for entry in entries}.values(), key=lambda item: item["releaseDate"])


def main() -> None:
    if not WORKBOOK.exists():
        raise SystemExit(f"Missing workbook: {WORKBOOK}")
    registry = json.loads(REGISTRY.read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(WORKBOOK, data_only=False, read_only=True)
    nsa = read_wide_sheet(wb, "Data_NSA")
    sa = read_wide_sheet(wb, "Data_SA")
    weights, weight_vintage, december_month = read_weights(wb)

    months = sorted(set(nsa) & set(sa))
    latest_months = [month for month in months if val(nsa, month, "CUUR0000SA0") is not None and val(sa, month, "CUSR0000SA0") is not None][-3:]
    history_months = months[-132:]
    heatmap_months = months[-24:]
    ref_month = latest_months[-1]

    all_dec = val(nsa, december_month or "", "CUUR0000SA0")
    all_ref = val(nsa, ref_month, "CUUR0000SA0")

    entries: list[dict[str, Any]] = []
    for item in registry:
        code = item["item_code"]
        series_nsa = item.get("series_nsa")
        series_sa = item.get("series_sa")
        weight = weights.get(code, {})
        dec_ri = weight.get("decRi")
        comp_dec = val(nsa, december_month or "", series_nsa)
        proxy_sa_index, proxy_sa_mm = ({}, {})
        sa_method = "official" if series_sa else "none"
        if not series_sa and series_nsa:
            proxy_months = sorted(nsa)
            proxy_sa_index, proxy_sa_mm = derived_sa_proxy(nsa, proxy_months, series_nsa)
            if any(proxy_sa_mm.get(month) is not None for month in history_months):
                sa_method = "derived_nsa_seasonal_proxy"

        history = []
        for month in history_months:
            nsa_index = val(nsa, month, series_nsa)
            prior = add_months(month, -1)
            official_sa_index = val(sa, month, series_sa)
            official_prior_sa_index = val(sa, prior, series_sa)
            sa_index = official_sa_index if official_sa_index is not None else proxy_sa_index.get(month)
            prior_sa_index = official_prior_sa_index if official_prior_sa_index is not None else proxy_sa_index.get(prior)
            sa_mm = pct_change(official_sa_index, official_prior_sa_index)
            if sa_mm is None and official_sa_index is None:
                sa_mm = proxy_sa_mm.get(month)
            prior_year = add_months(month, -12)
            prior_ri = price_updated_ri(
                dec_ri,
                val(nsa, prior, series_nsa),
                comp_dec,
                val(nsa, prior, "CUUR0000SA0"),
                all_dec,
            )
            history.append(
                {
                    "month": month,
                    "nsaIndex": nsa_index,
                    "saIndex": sa_index,
                    "saMm": sa_mm,
                    "nsaYoy": pct_change(nsa_index, val(nsa, prior_year, series_nsa)),
                    "ri": price_updated_ri(dec_ri, nsa_index, comp_dec, val(nsa, month, "CUUR0000SA0"), all_dec),
                    "contribution": contribution_pp(prior_ri, sa_index, prior_sa_index),
                    "saMethod": sa_method,
                }
            )

        latest = next((point for point in history if point["month"] == ref_month), history[-1])
        current_ri = price_updated_ri(dec_ri, val(nsa, ref_month, series_nsa), comp_dec, all_ref, all_dec)
        entries.append(
            {
                "itemCode": code,
                "name": item["name"],
                "parent": item.get("parent"),
                "level": item.get("level", item.get("display_level", 0)),
                "displayLevel": item.get("display_level", 0),
                "isItemStratum": bool(item.get("is_item_stratum")),
                "formula": item.get("formula", ""),
                "collection": item.get("collection", ""),
                "qaMethod": item.get("qa_method", ""),
                "altData": item.get("alt_data", ""),
                "notes": item.get("notes", ""),
                "links": item.get("links", []),
                "seriesNsa": series_nsa,
                "seriesSa": series_sa,
                "saMethod": sa_method,
                "decRi": dec_ri,
                "currentRi": current_ri,
                "weightVintage": weight.get("weightVintage"),
                "decemberMonth": december_month,
                "latest": latest,
                "history": history,
            }
        )

    by_code = {entry["itemCode"]: entry for entry in entries}
    headline = by_code["SA0"]
    core = by_code.get("SA0L1E")
    top_contributors = sorted(
        [
            {
                "itemCode": entry["itemCode"],
                "name": entry["name"],
                "contribution": entry["latest"].get("contribution"),
                "currentRi": entry.get("currentRi"),
                "saMm": entry["latest"].get("saMm"),
            }
            for entry in entries
            if entry["latest"].get("contribution") is not None
            and entry.get("formula") != "aggregate"
            and (entry.get("currentRi") or 0) > 0
        ],
        key=lambda row: abs(row["contribution"] or 0),
        reverse=True,
    )[:25]

    release_calendar = parse_release_calendar()
    next_release = next((entry for entry in release_calendar if entry["releaseDate"] >= date.today().isoformat()), release_calendar[-1])

    payload = {
        "generatedAt": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        "refMonth": ref_month,
        "latestMonths": latest_months,
        "historyMonths": history_months,
        "heatmapMonths": heatmap_months,
        "releaseDate": next_release["releaseDate"],
        "nextRelease": next_release,
        "source": "U.S. Bureau of Labor Statistics",
        "weightVintage": weight_vintage,
        "decemberMonth": december_month,
        "headline": {
            "itemCode": "SA0",
            "saMm": headline["latest"].get("saMm"),
            "nsaYoy": headline["latest"].get("nsaYoy"),
            "saIndex": headline["latest"].get("saIndex"),
            "nsaIndex": headline["latest"].get("nsaIndex"),
        },
        "core": {
            "itemCode": "SA0L1E",
            "saMm": core["latest"].get("saMm") if core else None,
            "nsaYoy": core["latest"].get("nsaYoy") if core else None,
            "saIndex": core["latest"].get("saIndex") if core else None,
            "nsaIndex": core["latest"].get("nsaIndex") if core else None,
        },
        "topContributors": top_contributors,
        "entries": entries,
    }
    OUTPUT.write_text(json.dumps(payload, indent=2, allow_nan=False), encoding="utf-8")
    CALENDAR_OUTPUT.write_text(json.dumps(release_calendar, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT.relative_to(ROOT)} with {len(entries)} entries.")


if __name__ == "__main__":
    main()
